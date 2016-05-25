from openpyxl import Workbook, load_workbook
from os import listdir
import os
import csv

class Parser:

	path_to_raw_files 	 = 'raw_files'
	path_to_parsed_files = 'parsed'

	def __init__(self):
		files = self.find_csv_filenames(Parser.path_to_raw_files)
		for path in files:
			path_name = path
			path 	  = '%s/%s' % (Parser.path_to_raw_files,path)
			self.parse_data(path, path_name)

	def find_csv_filenames( self, path_to_raw_files, suffix='.csv'):
		filenames = listdir(path_to_raw_files)
		return [ filename for filename in filenames if filename.endswith( suffix ) ]

	def parse_data(self, path, path_name):
	    ctr = 0
	    with open(path, 'rb') as csvfile:
	        spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
	        data = {}

	        for row in spamreader:
	            ctr += 1
	            res = row[0].split(";")
	            res = [r.replace('"','') for r in res]

	            if ctr == 1:
	                header = res
	            else:
	                sensor_id    = res[0]
	                device_sn_id = res[1]
	                sensor_code  = res[2]

	                format_res = {}

	                if device_sn_id in data:
	                    if sensor_code in data[device_sn_id]:
	                        format_res[len(data[device_sn_id][sensor_code]) + 1] = res
	                        data[device_sn_id][sensor_code].update(format_res)
	                        # print ''
	                        pass
	                    else:
	                        data[device_sn_id][sensor_code] = {}
	                        format_res[len(data[device_sn_id][sensor_code]) + 1] = res
	                        data[device_sn_id][sensor_code] = format_res
	                else:
	                    data[device_sn_id] = {}
	                    format_res[1] = res
	                    data[device_sn_id][sensor_code] = format_res

	            # if ctr == 10000:
	            #     break


	        # print '{} Recods'.format(ctr)
	        self.create_file(data, header, path_name)

	def create_file(self, data, header, path_name):
	    path_name 	= path_name.split('.')
	    devices 	= [row for row in data.items()]
	    ctr 		= 1

	    for key,val in devices:
	        device_id = key

	        wb = Workbook()
	        del wb['Sheet']

	        for k,v in val.items():
	            sensor_code = k
	            row_data    = v

	            ws = wb.create_sheet(sensor_code)
	            ws.append(header)
	            ws.title = sensor_code

	            for k1,v1 in row_data.iteritems():
	                ctr += 1
	                ws.append(v1)
	                print '{}. {}'.format(ctr,v1)

	        parsed_dir = 'parsed/%s' % (path_name[0])


	        if not os.path.exists(parsed_dir):
	        	os.makedirs(parsed_dir)


	        wb.save('parsed/%s/%s.xlsx' % (path_name[0],device_id) )


#----------------------------------------------------------------------
if __name__ == "__main__":
    Parser()
    
